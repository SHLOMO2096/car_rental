from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.crud.customer import (
    crud_customer,
    normalize_email,
    normalize_id_number,
    normalize_name,
    normalize_phone,
)

FIELD_ALIASES = {
    "name": {
        "name", "customer", "customer name", "client", "client name",
        "שם", "שם לקוח", "שם חשבון", "לקוח", "חשבון",
    },
    "address": {
        "address", "street", "location", "city", "addr",
        "כתובת", "מען", "עיר", "יישוב", "רחוב",
    },
    "phone": {
        "phone", "mobile", "cell", "telephone", "phones", "tel", "contact phone",
        "טלפון", "טלפונים", "נייד", "פלאפון", "טל", "מספר טלפון",
    },
    "email": {
        "email", "e-mail", "mail", "email address",
        "מייל", "דואל", 'דוא"ל', "אימייל", "דואר אלקטרוני",
    },
    "id_number": {
        "id", "id number", "id_no", "identity", "customer id", "company id", "vat", "tax id",
        "תעודת זהות", "תז", "מספר זהות", "מספר זיהוי", "עוסק מורשה", "ח.פ", "חפ",
    },
}


def normalize_header(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    for ch in ["\n", "\r", "\t", "_", "-", "/", "\\", ".", ",", ":", ";", '"', "'", "(", ")"]:
        text = text.replace(ch, " ")
    return " ".join(text.split())


def _clean_value(value) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned or cleaned == "-":
        return None
    return cleaned


def _pick_first_phone(raw: str | None) -> str | None:
    if not raw:
        return None
    pieces = [p for p in raw.replace("\t", " ").split(" ") if p.strip()]
    for piece in pieces:
        normalized = normalize_phone(piece)
        if normalized and len(normalized) >= 7:
            return normalized
    return normalize_phone(raw)


def _match_field(header_value: str | None) -> str | None:
    normalized = normalize_header(header_value)
    if not normalized:
        return None
    for field_name, aliases in FIELD_ALIASES.items():
        if normalized in aliases:
            return field_name
    return None


def _detect_header_row(rows: list[list]) -> tuple[int, dict[str, int]]:
    best_idx = -1
    best_score = -1
    best_mapping: dict[str, int] = {}

    for idx, row in enumerate(rows[:12]):
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            matched_field = _match_field(cell)
            if matched_field and matched_field not in mapping:
                mapping[matched_field] = col_idx
        score = len(mapping)
        if score > best_score:
            best_idx = idx
            best_score = score
            best_mapping = mapping

    if best_score <= 0 or "name" not in best_mapping:
        raise ValueError("לא נמצאה שורת כותרות מתאימה. צריך לפחות עמודת שם לקוח.")
    return best_idx, best_mapping


def _rows_from_csv_bytes(content: bytes) -> list[list[str]]:
    for encoding in ("utf-8-sig", "cp1255", "utf-8"):
        try:
            text = content.decode(encoding)
            return list(csv.reader(StringIO(text)))
        except UnicodeDecodeError:
            continue
    raise ValueError("לא ניתן לקרוא את קובץ ה-CSV. מומלץ לשמור כ-UTF-8 או Excel.")


def _rows_from_excel_bytes(content: bytes) -> list[list[str]]:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def load_rows_from_bytes(filename: str, content: bytes) -> list[list]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _rows_from_csv_bytes(content)
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
        return _rows_from_excel_bytes(content)
    if lower_name.endswith(".xls"):
        raise ValueError("קובץ Excel ישן (.xls) אינו נתמך כרגע. שמור אותו כ-.xlsx ונסה שוב.")
    raise ValueError("נתמך ייבוא מקבצי CSV או Excel בלבד (.csv, .xlsx, .xlsm)")


def load_rows_from_path(path: Path) -> list[list]:
    suffix = path.suffix.lower()
    content = path.read_bytes()
    return load_rows_from_bytes(path.name if suffix else "customers.csv", content)


def parse_customer_rows(raw_rows: list[list]) -> list[dict]:
    rows, _, _ = parse_customer_rows_with_report(raw_rows)
    return rows


def parse_customer_rows_with_report(raw_rows: list[list]) -> tuple[list[dict], list[dict], int]:
    if not raw_rows:
        return [], [], 0

    header_idx, mapping = _detect_header_row(raw_rows)
    body = raw_rows[header_idx + 1 :]
    dedup: dict[tuple, dict] = {}
    issues: list[dict] = []
    skipped = 0

    for idx, row in enumerate(body, start=header_idx + 2):
        if not row or not any(_clean_value(cell) for cell in row):
            continue

        name = _clean_value(row[mapping["name"]] if mapping.get("name") is not None and len(row) > mapping["name"] else None)
        if not name:
            skipped += 1
            issues.append({"row": idx, "level": "error", "field": "name", "message": "שורה ללא שם לקוח דולגה"})
            continue

        address = _clean_value(row[mapping["address"]] if mapping.get("address") is not None and len(row) > mapping["address"] else None)
        raw_phone = _clean_value(row[mapping["phone"]] if mapping.get("phone") is not None and len(row) > mapping["phone"] else None)
        raw_email = _clean_value(row[mapping["email"]] if mapping.get("email") is not None and len(row) > mapping["email"] else None)
        phone = _pick_first_phone(raw_phone)
        email = normalize_email(raw_email)
        id_number = normalize_id_number(_clean_value(row[mapping["id_number"]] if mapping.get("id_number") is not None and len(row) > mapping["id_number"] else None))

        if raw_phone and not phone:
            issues.append({"row": idx, "level": "warning", "field": "phone", "message": "מספר טלפון לא תקין נשמר כריק"})
        if email and "@" not in email:
            issues.append({"row": idx, "level": "warning", "field": "email", "message": "מייל לא תקין נשמר כריק"})
            email = None

        name_norm = normalize_name(name)
        addr_norm = normalize_name(address)
        if email:
            key = ("email", email)
        elif phone:
            key = ("phone", phone)
        elif id_number:
            key = ("id_number", id_number)
        else:
            key = ("name_addr", name_norm, addr_norm)

        current = dedup.get(key)
        if current is None:
            dedup[key] = {
                "name": name.strip(),
                "address": address,
                "phone": phone,
                "email": email,
                "id_number": id_number,
            }
            continue

        current["address"] = current["address"] or address
        current["phone"] = current["phone"] or phone
        current["email"] = current["email"] or email
        current["id_number"] = current["id_number"] or id_number

    return sorted(dedup.values(), key=lambda r: normalize_name(r["name"])), issues, skipped


def write_clean_csv(target_path: Path, rows: list[dict]) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with target_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["name", "address", "phone", "email", "id_number"])
        writer.writeheader()
        writer.writerows(rows)


def import_customers_to_db(db, rows: Iterable[dict], *, issues: list[dict] | None = None, skipped: int = 0) -> dict:
    from app.models.customer import Customer

    inserted = 0
    touched = 0
    known_ids = {row[0] for row in db.query(Customer.id).all()}

    for row in rows:
        customer = crud_customer.upsert_contact(
            db,
            name=row["name"],
            address=row.get("address"),
            phone=row.get("phone"),
            email=row.get("email"),
            id_number=row.get("id_number"),
        )
        if customer:
            touched += 1
        if customer and customer.id not in known_ids:
            inserted += 1
            known_ids.add(customer.id)
    db.commit()
    return {
        "processed": touched,
        "inserted": inserted,
        "updated": max(touched - inserted, 0),
        "skipped": skipped,
        "issues": issues or [],
    }

